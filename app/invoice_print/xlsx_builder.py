# app/invoice_print/xlsx_builder.py
# ============================================================
# Печатная форма счёта-фактуры в Excel (.xlsx) — тот же макет, что и
# pdf_builder.py (рамка "Увага!" сверху, блок Постачальник/Одержувач/
# Платник, таблица позиций с итогами тремя последними строками,
# сумма прописью). Запрошено Вахтангом 2026-08-31 сразу после
# PDF-версии ("нужно ещё экселевский вариант счёта").
#
# Данные — тот же build_invoice_print_data() из data.py, что и у PDF
# (см. комментарий в data.py: расчёт сумм/прописи/реквизитов не
# дублируется между builder'ами). Здесь только вёрстка под xlsx.
#
# Библиотека — openpyxl: чистый Python, без системных зависимостей
# (тот же критерий, что привёл к reportlab вместо WeasyPrint для PDF,
# см. pdf_builder.py) — ставится обычным pip и одинаково работает в
# SQLite-тестах и на Railway.
# ============================================================

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from app.invoice_print.data import InvoicePrintData
from app.invoice_print.amount_in_words import amount_in_words_uk

_THIN = Side(style="thin", color="000000")
_BORDER_ALL = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_BOLD = Font(bold=True)
_BOLD_12 = Font(bold=True, size=12)
_BOLD_10 = Font(bold=True, size=10)

# Число колонок таблицы позиций (совпадает с header в pdf_builder.py)
_N_COLS = 6


def build_invoice_xlsx(data: InvoicePrintData) -> bytes:
    """Возвращает готовый xlsx (bytes) печатной формы счёта."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Рахунок-фактура"

    for col, width in zip("ABCDEF", [5, 40, 8, 12, 16, 16]):
        ws.column_dimensions[col].width = width

    row = 1

    # --- "Увага!" в рамке ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    cell = ws.cell(row=row, column=1, value="Увага!")
    cell.font = _BOLD_12
    cell.alignment = Alignment(horizontal="center")
    row += 1

    warning_lines = [
        "У зв'язку з відкриттям нового розрахункового рахунку, реквізити змінені.",
        "Рахунок закривається актом виконаних робіт.",
    ]
    for line in warning_lines:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
        cell = ws.cell(row=row, column=1, value=line)
        cell.font = _BOLD
        cell.alignment = Alignment(horizontal="center")
        row += 1

    # Рамка вокруг блока "Увага!"
    for r in range(1, row):
        for c in range(1, _N_COLS + 1):
            ws.cell(row=r, column=c).border = _BORDER_ALL
    row += 1

    # --- Постачальник / Одержувач / Платник ---
    def label_value(label, value, bold_label=True):
        nonlocal row
        lbl_cell = ws.cell(row=row, column=1, value=label)
        if bold_label:
            lbl_cell.font = Font(bold=True, underline="single")
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        val_cell = ws.cell(row=row, column=2, value=value)
        val_cell.alignment = Alignment(wrap_text=True)
        row += 1

    label_value("Постачальник", data.firm_full_name)

    egrpou_line = f"ЄДРПОУ {data.firm_egrpou_code}" if data.firm_egrpou_code else ""
    if data.firm_phone:
        egrpou_line += f", тел. {data.firm_phone}."
    if egrpou_line:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2, value=egrpou_line)
        row += 1

    if data.firm_bank_account or data.firm_bank_name:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2, value=f"Р/р {data.firm_bank_account} в {data.firm_bank_name}")
        row += 1

    if data.firm_bank_mfo:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2, value=f"МФО {data.firm_bank_mfo}")
        row += 1

    if data.firm_tax_id or data.firm_vat_certificate_number:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2,
                value=f"ІПН {data.firm_tax_id}, номер свідоцтва {data.firm_vat_certificate_number}")
        row += 1

    if data.firm_address:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2, value=f"Адреса {data.firm_address}")
        row += 1

    label_value("Одержувач", data.client_full_name)
    if data.client_phone:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=_N_COLS)
        ws.cell(row=row, column=2, value=f"тел. {data.client_phone}")
        row += 1

    label_value("Платник", "той самий" if data.is_same_payer else data.client_full_name)

    row += 1

    # --- Заголовок документа ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    cell = ws.cell(row=row, column=1, value=f"Рахунок-фактура № {data.document_number}")
    cell.font = _BOLD_12
    cell.alignment = Alignment(horizontal="center")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    cell = ws.cell(row=row, column=1, value=f"від {data.document_date_str}")
    cell.alignment = Alignment(horizontal="center")
    row += 2

    # --- Таблица позиций ---
    header = ["№", "Назва", "Од.", "Кількість", "Ціна без ПДВ", "Сума без ПДВ"]
    header_row = row
    for col, text in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=col, value=text)
        cell.font = _BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER_ALL
    row += 1

    for line in data.lines:
        values = [line.position, line.name, line.unit,
                  line.quantity, line.unit_price, line.line_total]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _BORDER_ALL
            if col == 1:
                cell.alignment = Alignment(horizontal="center")
            elif col == 4:
                cell.number_format = "0.000"
                cell.alignment = Alignment(horizontal="right")
            elif col in (5, 6):
                cell.number_format = "0.00"
                cell.alignment = Alignment(horizontal="right")
        row += 1

    # --- Итоги (те же 3 строки, что в PDF) ---
    totals = [
        ("Разом без ПДВ:", data.total_excl_vat),
        ("ПДВ:", data.vat_amount),
        ("Всього з ПДВ:", data.total_incl_vat),
    ]
    for label, value in totals:
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        label_cell = ws.cell(row=row, column=2, value=label)
        label_cell.font = _BOLD
        label_cell.alignment = Alignment(horizontal="right")
        value_cell = ws.cell(row=row, column=6, value=value)
        value_cell.font = _BOLD
        value_cell.number_format = "0.00"
        value_cell.alignment = Alignment(horizontal="right")
        for col in range(1, _N_COLS + 1):
            ws.cell(row=row, column=col).border = _BORDER_ALL
        row += 1

    row += 1

    # --- Сумма прописью ---
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    ws.cell(row=row, column=1, value="Всього на суму:")
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    cell = ws.cell(row=row, column=1, value=amount_in_words_uk(data.total_incl_vat))
    cell.font = _BOLD_10
    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=_N_COLS)
    ws.cell(row=row, column=1, value=f"ПДВ:    {data.vat_amount:.2f} грн.")
    row += 2

    # --- Подписи ---
    ws.cell(row=row, column=1, value="Виписав(ла):")
    row += 2
    ws.cell(row=row, column=1, value="Рахунок дійсний до сплати до   .  .")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
